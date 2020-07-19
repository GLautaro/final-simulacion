from openpyxl import Workbook,load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def Truncate(n, decimals=0):
    multiplier = 10 ** decimals
    return int(n * multiplier) / multiplier


def GenerarExcel(dict_dataframes, nombre_archivo):
    wb = Workbook(nombre_archivo)
    for nombre,df in dict_dataframes.items():
        ws = wb.create_sheet(nombre)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    wb.save(nombre_archivo)


def crearColumnasParcialesDataFrame(cant_mesas):
    '''
    Crea una lista con las columnas correspondientes a los 
    datos siempre presentes en el dataframe
    '''
    eventos = [
        "Evento Actual",
        "Reloj",
        "Tiempo entre llegadas Cliente",
        "Próxima llegada Cliente",
        "RND_Decision",
        "Decision_cliente",
        "Tiempo compra ticket",
        "Fin compra ticket",
        "Tiempo entrega pedido emp. 1",
        "Fin entrega pedido emp. 1",
        "Tiempo entrega pedido emp. 2",
        "Fin entrega pedido emp. 2",
        "Tiempo uso mesa"
    ]

    mesas_evento = []
    for i in range(cant_mesas):
        mesas_evento.append("Fin uso mesa " + str(i+1))

    colas_estados = [
        "Cola dueño",
        "Estado dueño",
        "Cola entregas",
        "Estado Empleado 1",
        "Estado Empleado 2"
    ]

    mesas_estado = []
    for i in range(cant_mesas):
        mesas_estado.append("Estado Mesa " + str(i+1))
    
    estadisticos = [
        "Cantidad Clientes Finalizados",
        "Acumulador tiempo permanencia"
    ]

    return eventos + mesas_evento + colas_estados + mesas_estado + estadisticos